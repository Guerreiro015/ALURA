import os
os.system("cls" or None)

import openpyxl

# ABRIR UM ARQUIVO EXSITENTE

plan = openpyxl.load_workbook('dados.xlsx')

print(plan.sheetnames)

#selecionar aba que vai usar
planilha = plan.active
# ou escolhe a aba pelo nome
planilha = plan["NOMES"]

#pecrorrer as linhas
for linha in plan.active.iter_rows(max_row=1):
    #pecorrer as colunas
    for celula in linha:
        print(celula.value)
        
    #Alterar valor das celulas
for celula in plan.active["C"]:
    if celula.value == "Vassora Sabiá":
        linha = celula.row
        plan.active[f"C{linha}"] = "Varredor"
    print(celula.value)

tamanho = plan.active["a1:c20"]

for x,y,z in tamanho:
    print(x.value,y.value,z.value)



 
        
#plan.save("dados2.xlsx")
