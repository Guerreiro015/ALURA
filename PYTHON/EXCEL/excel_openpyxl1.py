
import os
os.system("cls" or None)

# import pandas as pd

# tabela = pd.read_excel("dados.xlsx")
# tabela.head()

###################################################
#==================================================
import openpyxl

#Criar uma planilha
tabela = openpyxl.Workbook()

#mostrar a aba da planilha
print(tabela.sheetnames)

#criar uma aba na planilha
tabela.create_sheet("frutas")

#selecionar uma aba na planilha
aba = tabela["frutas"]

#adicionar dados na planilha
aba.append(['Nome',"Quantidade",'preço'])
aba.append(['banana',2,2.9])
aba.append(['caju',23,4])
aba.append(['laranja',4,4.8])
aba.append(['melancia',10,2.3])
aba.append(['banana',2,2.9])

aba = tabela["Sheet"]
aba.append(['NOME',"QuanT.",'PRECO'])
aba.append(['ARROZ',2,2.9])
aba.append(['FEIJÃO',23,4])
aba.append(['laranja',4,4.8])
aba.append(['melancia',10,2.3])
aba.append(['banana',2,2.9])

print(tabela.sheetnames)

#salvar a planilha - Obs: o openpyxl só o formato xlxs
tabela.save("teste.xlsx")

# ABRIR UM ARQUIVO EXSITENTE

# plan = openpyxl.Workbook('teste.xlsx')

# print(plan.sheetnames)
# planilha = plan["frutas"]








